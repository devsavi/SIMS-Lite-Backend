"""
Product management endpoints -- Phase 2.

GET    /api/v1/products/                  -- list (paginated, search, filter)
POST   /api/v1/products/                  -- create (auto SKU + barcode)
GET    /api/v1/products/import-template   -- download Excel import template
POST   /api/v1/products/import            -- bulk import from Excel
GET    /api/v1/products/{id}              -- get product
PUT    /api/v1/products/{id}              -- update product
DELETE /api/v1/products/{id}              -- soft delete
POST   /api/v1/products/{id}/image        -- upload product image
DELETE /api/v1/products/{id}/image        -- delete product image
GET    /api/v1/products/{id}/barcode      -- get barcode PNG
"""

from __future__ import annotations

import uuid
from io import BytesIO

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_client_ip, get_current_user, require_permission
from app.core.exceptions import ValidationError as AppValidationError
from app.core.logging import get_logger
from app.database.engine import get_db
from app.models.user import User
from app.repositories.master_data import (
    BrandRepository,
    CategoryRepository,
    SupplierRepository,
    UnitOfMeasureRepository,
)
from app.schemas.base import PaginatedResponse, PaginationMeta, SuccessResponse
from app.schemas.master_data import (
    ProductCreate,
    ProductImportRow,
    ProductRead,
    ProductUpdate,
)
from app.services.master_data import ProductService
from app.services.report import ReportService

logger = get_logger(__name__)
router = APIRouter()

_ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
_MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5 MB


def _svc(db: AsyncSession = Depends(get_db)) -> ProductService:
    return ProductService(db)


def _report_svc() -> ReportService:
    return ReportService()


# ---------------------------------------------------------------------------
# List
# ---------------------------------------------------------------------------


@router.get("/", response_model=PaginatedResponse[ProductRead], summary="List products")
async def list_products(
    page: int = Query(default=1, ge=1),
    size: int = Query(default=20, ge=1, le=100),
    search: str | None = Query(default=None),
    active_only: bool = Query(default=False),
    category_id: uuid.UUID | None = Query(default=None),
    brand_id: uuid.UUID | None = Query(default=None),
    supplier_id: uuid.UUID | None = Query(default=None),
    current_user: User = Depends(get_current_user),
    svc: ProductService = Depends(_svc),
) -> PaginatedResponse[ProductRead]:
    items, total = await svc.list(
        page=page, size=size, search=search, active_only=active_only,
        category_id=category_id, brand_id=brand_id, supplier_id=supplier_id,
    )
    pages = (total + size - 1) // size if total else 0
    return PaginatedResponse(
        data=[ProductRead.model_validate(p) for p in items],
        pagination=PaginationMeta(page=page, size=size, total=total, pages=pages),
    )


# ---------------------------------------------------------------------------
# Import template  (must be before /{id} routes to avoid path conflict)
# ---------------------------------------------------------------------------


@router.get(
    "/import-template",
    summary="Download Excel product import template",
    response_class=Response,
)
async def download_import_template(
    current_user: User = Depends(get_current_user),
    report_svc: ReportService = Depends(_report_svc),
) -> Response:
    data = report_svc.generate_product_import_template()
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"},
    )


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------


@router.post(
    "/import",
    response_model=SuccessResponse[dict],
    status_code=status.HTTP_201_CREATED,
    summary="Bulk import products from Excel",
)
async def import_products(
    file: UploadFile = File(...),
    current_user: User = Depends(require_permission("master_data:write")),
    db: AsyncSession = Depends(get_db),
    ip: str | None = Depends(get_client_ip),
) -> SuccessResponse[dict]:
    if file.content_type not in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    file_bytes = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {exc}")

    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

    # Resolve lookup tables
    cats = {c.name.lower(): c for c in (await CategoryRepository(db).get_all())}
    brands = {b.name.lower(): b for b in (await BrandRepository(db).get_all())}
    uoms = {u.symbol.lower(): u for u in (await UnitOfMeasureRepository(db).get_all())}
    suppliers = {s.supplier_code.lower(): s for s in (await SupplierRepository(db).get_all())}

    svc = ProductService(db)
    created = 0
    errors: list[dict] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        if not any(row_data.values()):
            continue  # skip empty rows

        name = row_data.get("name*") or row_data.get("name")
        if not name:
            errors.append({"row": row_idx, "error": "name is required"})
            continue

        cat = cats.get((row_data.get("category_name") or "").lower())
        brand = brands.get((row_data.get("brand_name") or "").lower())
        uom = uoms.get((row_data.get("uom_symbol") or "").lower())
        supplier = suppliers.get((row_data.get("supplier_code") or "").lower())

        try:
            cost = float(row_data.get("cost_price")) if row_data.get("cost_price") else None
            sell = float(row_data.get("selling_price")) if row_data.get("selling_price") else None
            payload = ProductCreate(
                name=str(name),
                description=str(row_data.get("description") or "") or None,
                category_id=cat.id if cat else None,
                brand_id=brand.id if brand else None,
                uom_id=uom.id if uom else None,
                supplier_id=supplier.id if supplier else None,
                cost_price=cost,
                selling_price=sell,
                reorder_level=int(row_data.get("reorder_level") or 0),
                reorder_quantity=int(row_data.get("reorder_quantity") or 0),
            )
            await svc.create(payload, actor=current_user, ip_address=ip)
            created += 1
        except Exception as exc:
            errors.append({"row": row_idx, "error": str(exc)})

    return SuccessResponse(data={"created": created, "errors": errors})


# ---------------------------------------------------------------------------
# CRUD
# ---------------------------------------------------------------------------


@router.post(
    "/",
    response_model=SuccessResponse[ProductRead],
    status_code=status.HTTP_201_CREATED,
    summary="Create product",
)
async def create_product(
    payload: ProductCreate,
    current_user: User = Depends(require_permission("master_data:write")),
    svc: ProductService = Depends(_svc),
    ip: str | None = Depends(get_client_ip),
) -> SuccessResponse[ProductRead]:
    product = await svc.create(payload, actor=current_user, ip_address=ip)
    return SuccessResponse(data=ProductRead.model_validate(product))


@router.get("/{product_id}", response_model=SuccessResponse[ProductRead], summary="Get product")
async def get_product(
    product_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    svc: ProductService = Depends(_svc),
) -> SuccessResponse[ProductRead]:
    product = await svc.get(product_id)
    return SuccessResponse(data=ProductRead.model_validate(product))


@router.put("/{product_id}", response_model=SuccessResponse[ProductRead], summary="Update product")
async def update_product(
    product_id: uuid.UUID,
    payload: ProductUpdate,
    current_user: User = Depends(require_permission("master_data:write")),
    svc: ProductService = Depends(_svc),
    ip: str | None = Depends(get_client_ip),
) -> SuccessResponse[ProductRead]:
    product = await svc.update(product_id, payload, actor=current_user, ip_address=ip)
    return SuccessResponse(data=ProductRead.model_validate(product))


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT, summary="Delete product")
async def delete_product(
    product_id: uuid.UUID,
    current_user: User = Depends(require_permission("master_data:delete")),
    svc: ProductService = Depends(_svc),
    ip: str | None = Depends(get_client_ip),
):
    await svc.delete(product_id, actor=current_user, ip_address=ip)


# ---------------------------------------------------------------------------
# Image endpoints
# ---------------------------------------------------------------------------


@router.post(
    "/{product_id}/image",
    response_model=SuccessResponse[ProductRead],
    summary="Upload product image",
)
async def upload_product_image(
    product_id: uuid.UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(require_permission("master_data:write")),
    svc: ProductService = Depends(_svc),
    ip: str | None = Depends(get_client_ip),
) -> SuccessResponse[ProductRead]:
    content_type = file.content_type or ""
    if content_type not in _ALLOWED_IMAGE_TYPES:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported image type. Allowed: {', '.join(_ALLOWED_IMAGE_TYPES)}",
        )
    data = await file.read()
    if len(data) > _MAX_IMAGE_SIZE:
        raise HTTPException(status_code=400, detail="Image exceeds 5 MB limit.")
    product = await svc.upload_image(
        product_id, data, content_type, actor=current_user, ip_address=ip
    )
    return SuccessResponse(data=ProductRead.model_validate(product))


@router.delete(
    "/{product_id}/image",
    response_model=SuccessResponse[ProductRead],
    summary="Delete product image",
)
async def delete_product_image(
    product_id: uuid.UUID,
    current_user: User = Depends(require_permission("master_data:write")),
    svc: ProductService = Depends(_svc),
    ip: str | None = Depends(get_client_ip),
) -> SuccessResponse[ProductRead]:
    product = await svc.delete_image(product_id, actor=current_user, ip_address=ip)
    return SuccessResponse(data=ProductRead.model_validate(product))


# ---------------------------------------------------------------------------
# Barcode endpoint
# ---------------------------------------------------------------------------


@router.get(
    "/{product_id}/barcode",
    summary="Get product barcode as PNG",
    response_class=Response,
)
async def get_product_barcode(
    product_id: uuid.UUID,
    current_user: User = Depends(get_current_user),
    svc: ProductService = Depends(_svc),
) -> Response:
    product = await svc.get(product_id)
    png_bytes = svc.generate_barcode_png(product.barcode)
    return Response(
        content=png_bytes,
        media_type="image/png",
        headers={
            "Content-Disposition": f"inline; filename=barcode_{product.sku}.png",
            "X-Barcode-Value": product.barcode,
            "X-Product-SKU": product.sku,
        },
    )

