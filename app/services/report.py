"""
Report generation service -- Phase 2.

Generates Excel workbooks for product, supplier, and category reports.
Uses openpyxl for Excel export.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from app.core.logging import get_logger
from app.models.master_data import Category, Product, Supplier

logger = get_logger(__name__)

# --- Column styles for openpyxl ---
HEADER_FILL_COLOR = "2B5797"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_COLOR = "EBF0F8"


def _make_workbook():
    """Create and return an openpyxl Workbook with styles."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    return wb


def _style_header_row(ws, num_cols: int) -> None:
    """Apply header styling to the first row of a worksheet."""
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_rows(ws, total_rows: int, num_cols: int) -> None:
    """Apply alternating row colours to data rows."""
    from openpyxl.styles import Alignment, PatternFill

    alt_fill = PatternFill(fill_type="solid", fgColor=ALT_ROW_COLOR)

    for row in range(2, total_rows + 2):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if row % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")


def _auto_fit_columns(ws) -> None:
    """Auto-size columns based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


class ReportService:
    """Generates Excel reports for master data entities."""

    def generate_product_report(self, products: list[Product]) -> bytes:
        """Build an Excel product report and return the bytes."""
        wb = _make_workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "SKU", "Barcode", "Name", "Category", "Brand",
            "UoM", "Supplier", "Cost Price", "Selling Price",
            "Reorder Level", "Reorder Qty", "Status", "Created At",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        for p in products:
            ws.append([
                p.sku,
                p.barcode,
                p.name,
                p.category.name if p.category else "",
                p.brand.name if p.brand else "",
                p.uom.symbol if p.uom else "",
                p.supplier.name if p.supplier else "",
                float(p.cost_price) if p.cost_price is not None else "",
                float(p.selling_price) if p.selling_price is not None else "",
                p.reorder_level,
                p.reorder_quantity,
                "Active" if p.is_active else "Inactive",
                p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            ])

        _style_data_rows(ws, len(products), len(headers))
        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_supplier_report(self, suppliers: list[Supplier]) -> bytes:
        """Build an Excel supplier report and return the bytes."""
        wb = _make_workbook()
        ws = wb.active
        ws.title = "Suppliers"

        headers = [
            "Supplier Code", "Name", "Contact Person", "Email",
            "Phone", "City", "Country", "Payment Terms", "Status", "Created At",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        for s in suppliers:
            ws.append([
                s.supplier_code,
                s.name,
                s.contact_person or "",
                s.email or "",
                s.phone or "",
                s.city or "",
                s.country or "",
                s.payment_terms or "",
                "Active" if s.is_active else "Inactive",
                s.created_at.strftime("%Y-%m-%d %H:%M") if s.created_at else "",
            ])

        _style_data_rows(ws, len(suppliers), len(headers))
        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_category_report(self, categories: list[Category]) -> bytes:
        """Build an Excel category report and return the bytes."""
        wb = _make_workbook()
        ws = wb.active
        ws.title = "Categories"

        headers = [
            "Name", "Slug", "Description", "Parent ID", "Status", "Created At",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        for c in categories:
            ws.append([
                c.name,
                c.slug,
                c.description or "",
                str(c.parent_id) if c.parent_id else "",
                "Active" if c.is_active else "Inactive",
                c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
            ])

        _style_data_rows(ws, len(categories), len(headers))
        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_product_import_template(self) -> bytes:
        """Return an Excel template for bulk product import."""
        wb = _make_workbook()
        ws = wb.active
        ws.title = "Product Import Template"

        headers = [
            "name*", "description", "category_name",
            "brand_name", "uom_symbol", "supplier_code",
            "cost_price", "selling_price",
            "reorder_level", "reorder_quantity",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        # Example row
        ws.append([
            "Example Product", "Product description", "Electronics",
            "BrandName", "pcs", "SUP-00001",
            10.50, 15.00, 5, 10,
        ])

        _auto_fit_columns(ws)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
